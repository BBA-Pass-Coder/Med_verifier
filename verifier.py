"""
Core medical creator verification logic.
Screenshots TikTok profiles and uses OpenAI Vision to check for medical attire.
"""

import asyncio
import base64
import csv
import json
import os
import random
import urllib.request
from pathlib import Path

OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")

SCREENSHOT_DIR = Path("/tmp/screenshots")
SCREENSHOT_DIR.mkdir(exist_ok=True)

USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
]

VISION_PROMPT = """You are analyzing a TikTok profile screenshot to determine if this creator is likely a real medical professional (doctor, nurse, pharmacist, dentist, PA, NP, etc.).

Look for these visual signals:

STRONG indicators (high confidence):
- Wearing scrubs (any color)
- Wearing a white lab coat
- Stethoscope visible (around neck or in hand)
- Hospital/clinic badge visible
- Clinical setting in background (exam room, hospital hallway, pharmacy shelves)
- Medical equipment visible

MODERATE indicators:
- Professional headshot style common in healthcare
- Anatomy posters or medical charts in background
- Prescription bottles or pharmaceutical products in professional context

WEAK/NEGATIVE indicators:
- No medical attire or setting visible
- Lifestyle/fitness influencer aesthetic
- Home setting with no medical context
- Supplement bottles without clinical framing

Respond in this exact JSON format and nothing else:
{
  "is_medical": true or false,
  "confidence": "high" or "medium" or "low",
  "signals_found": ["list of specific visual signals you spotted"],
  "likely_role": "doctor/nurse/pharmacist/dentist/unknown/not_medical",
  "reasoning": "one sentence explanation"
}"""

RECYCLE_EVERY = 15
DELAY_MIN = 3
DELAY_MAX = 8

# Global progress tracker
progress = {
    "status": "idle",  # idle, running, done, error
    "total": 0,
    "processed": 0,
    "current_handle": "",
    "medical_count": 0,
    "error_count": 0,
    "results": [],
}


def reset_progress():
    progress["status"] = "idle"
    progress["total"] = 0
    progress["processed"] = 0
    progress["current_handle"] = ""
    progress["medical_count"] = 0
    progress["error_count"] = 0
    progress["results"] = []


def load_handles_from_file(path: str) -> list[str]:
    target_cols = ["handle", "username", "tiktok_handle", "creator_handle"]
    handles = []

    if path.endswith((".xlsx", ".xlsm", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_idx = None
        for tc in target_cols:
            if tc in headers:
                col_idx = headers.index(tc)
                break
        if col_idx is None:
            return []
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[col_idx]
            if val:
                handles.append(str(val).strip())
        wb.close()
    else:
        with open(path, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                for col in ["handle", "username", "tiktok_handle", "creator_handle", "Handle", "Username"]:
                    if col in row and row[col]:
                        handles.append(row[col].strip())
                        break
    return handles


async def take_screenshot(page, handle: str) -> str | None:
    clean = handle.lstrip("@")
    url = f"https://www.tiktok.com/@{clean}"
    path = SCREENSHOT_DIR / f"{clean}.png"

    for attempt in range(3):
        try:
            await page.goto(url, wait_until="networkidle", timeout=30000)
            await page.wait_for_timeout(3000)

            for selector in [
                'button:has-text("Accept")',
                'button:has-text("Decline")',
                '[data-e2e="modal-close-inner-button"]',
                'button:has-text("Not now")',
            ]:
                try:
                    btn = page.locator(selector).first
                    if await btn.is_visible(timeout=1000):
                        await btn.click()
                        await page.wait_for_timeout(500)
                except Exception:
                    pass

            await page.evaluate("window.scrollTo(0, 0)")
            await page.wait_for_timeout(500)
            await page.screenshot(path=str(path), full_page=False)
            return str(path)

        except Exception as e:
            if attempt < 2:
                await page.wait_for_timeout(2000)
            else:
                return None
    return None


def analyze_screenshot(image_path: str) -> dict:
    with open(image_path, "rb") as f:
        img_data = base64.standard_b64encode(f.read()).decode("utf-8")

    payload = json.dumps({
        "model": "gpt-4o-mini",
        "max_tokens": 500,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{img_data}",
                            "detail": "low",
                        },
                    },
                    {"type": "text", "text": VISION_PROMPT},
                ],
            }
        ],
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=payload,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {OPENAI_API_KEY}",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read().decode())

        text = data["choices"][0]["message"]["content"].strip()
        if text.startswith("```"):
            text = text.split("\n", 1)[1].rsplit("```", 1)[0]

        return json.loads(text)

    except Exception as e:
        return {
            "is_medical": None,
            "confidence": "error",
            "signals_found": [],
            "likely_role": "error",
            "reasoning": f"API error: {e}",
        }


async def run_verification(handles: list[str]):
    from playwright.async_api import async_playwright

    reset_progress()
    progress["status"] = "running"
    progress["total"] = len(handles)

    try:
        async with async_playwright() as p:

            async def make_browser():
                ua = random.choice(USER_AGENTS)
                browser = await p.chromium.launch(
                    headless=True,
                    args=["--disable-blink-features=AutomationControlled", "--no-sandbox", "--disable-dev-shm-usage"],
                )
                context = await browser.new_context(
                    viewport={"width": random.choice([1280, 1366, 1440]), "height": random.choice([800, 900, 960])},
                    user_agent=ua,
                    locale=random.choice(["en-US", "en-GB"]),
                    timezone_id=random.choice(["America/New_York", "America/Chicago", "America/Los_Angeles"]),
                )
                await context.add_init_script("""
                    Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                    Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
                    window.chrome = { runtime: {} };
                """)
                page = await context.new_page()
                return browser, page

            browser, page = await make_browser()
            browser_count = 0

            for i, handle in enumerate(handles):
                clean = handle.lstrip("@")
                progress["current_handle"] = f"@{clean}"
                progress["processed"] = i

                # Recycle browser
                if browser_count >= RECYCLE_EVERY:
                    await browser.close()
                    browser, page = await make_browser()
                    browser_count = 0

                # Screenshot
                img_path = await take_screenshot(page, handle)

                if not img_path:
                    row = {
                        "handle": f"@{clean}",
                        "is_medical": None,
                        "confidence": "error",
                        "likely_role": "error",
                        "signals_found": "",
                        "reasoning": "Screenshot failed",
                    }
                    progress["results"].append(row)
                    progress["error_count"] += 1
                    continue

                # Analyze
                analysis = analyze_screenshot(img_path)
                row = {
                    "handle": f"@{clean}",
                    "is_medical": analysis.get("is_medical"),
                    "confidence": analysis.get("confidence"),
                    "likely_role": analysis.get("likely_role"),
                    "signals_found": "; ".join(analysis.get("signals_found", [])),
                    "reasoning": analysis.get("reasoning"),
                }
                progress["results"].append(row)

                if analysis.get("is_medical"):
                    progress["medical_count"] += 1
                if analysis.get("confidence") == "error":
                    progress["error_count"] += 1

                browser_count += 1

                # Delay
                if i < len(handles) - 1:
                    await asyncio.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

            await browser.close()

        progress["processed"] = progress["total"]
        progress["status"] = "done"

    except Exception as e:
        progress["status"] = "error"
        progress["current_handle"] = f"Error: {e}"
